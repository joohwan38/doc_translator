# python_backend/excel_handler.py
import openpyxl
import os
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, Protection
from copy import copy
import logging
import shutil
import traceback
from typing import Dict, Any, Optional, Callable, List

from interfaces import AbsExcelProcessor, AbsTranslator, AbsOllamaService
import config


logger = logging.getLogger(__name__)

def copy_cell_style(source_cell: Cell, target_cell: Cell):
    """openpyxl 셀 스타일을 복사합니다."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

class ExcelHandler(AbsExcelProcessor):
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """엑셀 파일의 정보를 분석하여 반환합니다."""
        print(f"[DEBUG] excel_handler.py: get_file_info 진입. 파일 경로: {file_path}") # [!INFO] 디버깅 print 1
        info = {
            "sheet_count": 0,
            "translatable_cell_count": 0,
            "total_text_char_count": 0,
            "error": None
        }
        try:
            print("[DEBUG] excel_handler.py: openpyxl.load_workbook 호출 직전") # [!INFO] 디버깅 print 2
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_vba=False)
            print("[DEBUG] excel_handler.py: 워크북 로드 성공") # [!INFO] 디버깅 print 3

            info["sheet_count"] = len(workbook.sheetnames)

            total_chars = 0
            cell_count = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            total_chars += len(cell.value)
                            cell_count += 1

            info["translatable_cell_count"] = cell_count
            info["total_text_char_count"] = total_chars
            print(f"[DEBUG] excel_handler.py: 분석 완료. 시트: {info['sheet_count']}, 셀: {cell_count}, 문자수: {total_chars}") # [!INFO] 디버깅 print 4

        except Exception as e:
            print(f"[DEBUG] excel_handler.py: get_file_info의 try-except 블록에서 오류 발생!") # [!INFO] 디버깅 print 5
            print("--- TRACEBACK START ---")
            traceback.print_exc()
            print("--- TRACEBACK END ---")

            logger.error(f"엑셀 파일 정보 분석 중 오류: {e}", exc_info=True)
            info["error"] = str(e)

        return info

    def translate_workbook(self, file_path: str, output_path: str, translator: AbsTranslator,
                           src_lang_ui_name: str, tgt_lang_ui_name: str, model_name: str,
                           ollama_service: 'AbsOllamaService',
                           progress_callback: Optional[Callable[[Any, str, float, str], None]] = None,
                           stop_event: Optional[Any] = None) -> Optional[str]:
        temp_input_path = None
        workbook = None # Initialize workbook
        try:
            # 1. 번역할 텍스트 추출 (읽기 전용으로 먼저 시도)
            texts_to_translate = []
            cell_map = []
            merged_cell_ranges_map = {}
            sheet_name_original_order = [] # To maintain order and map translated names back
            
            try:
                # 드로잉 요소 문제 없는 대부분의 파일을 위해 읽기/쓰기 모드로 먼저 시도
                temp_input_path = os.path.join(os.path.dirname(output_path), f"temp_readwrite_{os.path.basename(file_path)}")
                shutil.copy2(file_path, temp_input_path)
                workbook = openpyxl.load_workbook(temp_input_path)
                is_read_only_fallback = False
            except KeyError as ke:
                if "There is no item named 'xl/drawings/NULL'" in str(ke):
                    logger.warning(f"드로잉 요소 문제로 일반 로드 실패. 읽기 전용 모드로 재시도합니다: {ke}")
                    if temp_input_path and os.path.exists(temp_input_path): os.remove(temp_input_path)
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    is_read_only_fallback = True
                else:
                    raise # 다른 KeyError는 다시 발생시킴
            except Exception as e: # Catch any other exception during initial load
                logger.error(f"엑셀 파일 초기 로드 중 예상치 못한 오류 발생: {e}", exc_info=True)
                workbook = None # Ensure workbook is None if an error occurs

            if workbook is None:
                logger.error(f"엑셀 파일 '{file_path}'을(를) 로드하는 데 실패했습니다. 파일이 손상되었거나 지원되지 않는 형식일 수 있습니다.")
                return None # Indicate failure

            # Add sheet names to translation queue
            for sheet_name in workbook.sheetnames:
                texts_to_translate.append(sheet_name)
                sheet_name_original_order.append(sheet_name)

            for sheet in workbook.worksheets:
                sheet_name = sheet.title
                # 읽기 전용 모드에서는 merged_cells가 제대로 동작하지 않을 수 있으므로 예외 처리
                try:
                    merged_cell_ranges_map[sheet_name] = [str(merged_range) for merged_range in sheet.merged_cells.ranges]
                except Exception:
                    merged_cell_ranges_map[sheet_name] = []

                merged_cells_in_sheet = set()
                for merged_range in merged_cell_ranges_map.get(sheet_name, []):
                    # iter_rows, iter_cols 등은 읽기 전용에서만 사용 가능할 수 있음
                    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(merged_range)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            merged_cells_in_sheet.add(openpyxl.utils.get_column_letter(col) + str(row))

                for row in sheet.iter_rows():
                    for cell in row:
                        if not hasattr(cell, 'coordinate'): continue # EmptyCell 방지

                        # 병합된 셀의 첫번째가 아니면 건너뛰기 (병합 정보가 정확할 때만 유효)
                        if cell.coordinate in merged_cells_in_sheet:
                            is_start_cell = False
                            for r in merged_cell_ranges_map.get(sheet_name, []):
                                if cell.coordinate == r.split(':')[0]:
                                    is_start_cell = True
                                    break
                            if not is_start_cell: continue

                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            texts_to_translate.append(cell.value)
                            cell_map.append((sheet_name, cell.coordinate))

            if not texts_to_translate:
                logger.info("엑셀 파일에 번역할 텍스트가 없습니다. 원본을 복사합니다.")
                shutil.copy2(file_path, output_path)
                return output_path

            # 2. 텍스트 일괄 번역
            if progress_callback: progress_callback("Excel", "status_task_translating_cells", 0, f"Translating {len(texts_to_translate)} cells...")
            translated_texts = translator.translate_texts(
                texts_to_translate, src_lang_ui_name, tgt_lang_ui_name, model_name, ollama_service, 
                stop_event=stop_event, progress_callback=progress_callback,
                base_location_key="status_key_excel_file", base_task_key="status_task_translating_cells"
            )

            if stop_event and stop_event.is_set(): logger.info("엑셀 번역 작업이 중단되었습니다.")
            if len(texts_to_translate) != len(translated_texts): raise ValueError("번역된 텍스트와 원본 텍스트의 개수가 불일치합니다.")

            # Separate translated sheet names from translated cell values
            num_sheets = len(sheet_name_original_order)
            translated_sheet_names = translated_texts[:num_sheets]
            translated_cell_texts = translated_texts[num_sheets:]

            translated_sheet_name_map = {sheet_name_original_order[i]: translated_sheet_names[i] for i in range(num_sheets)}
            translated_map = {cell_map[i]: translated_cell_texts[i] for i in range(len(translated_cell_texts))}

            # 3. 번역된 내용을 워크북에 쓰기
            if is_read_only_fallback:
                # 읽기 전용 모드였으면, 새 워크북을 만들어 내용을 복사 (재조립)
                logger.info("읽기 전용 폴백 모드: 새 워크북을 생성하여 내용을 재조립합니다.")
                new_workbook = openpyxl.Workbook()
                # 기본 시트 제거
                if "Sheet" in new_workbook.sheetnames and len(new_workbook.sheetnames) == 1:
                    new_workbook.remove(new_workbook.active)

                for sheet_name in workbook.sheetnames:
                    source_sheet = workbook[sheet_name]
                    translated_sheet_title = translated_sheet_name_map.get(sheet_name, sheet_name) # Get translated name, fallback to original
                    target_sheet = new_workbook.create_sheet(title=translated_sheet_title) # Use translated title

                    # 셀 값 및 스타일 복사
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            if not hasattr(cell, 'coordinate'): continue
                            new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                            if (sheet_name, cell.coordinate) in translated_map:
                                translated_text = translated_map[(sheet_name, cell.coordinate)]
                                if "오류:" not in translated_text:
                                    new_cell.value = translated_text
                                else:
                                    new_cell.value = cell.value # 오류 시 원본 값 유지
                            else:
                                new_cell.value = cell.value
                            copy_cell_style(cell, new_cell)
                    
                    # 컬럼 너비 복사
                    for col_idx, column_dimension in source_sheet.column_dimensions.items():
                        if column_dimension.width is not None:
                            target_sheet.column_dimensions[col_idx].width = column_dimension.width

                    # 병합 정보 복사
                    if sheet_name in merged_cell_ranges_map:
                        for merged_range in merged_cell_ranges_map[sheet_name]:
                            target_sheet.merge_cells(merged_range)
                
                if progress_callback: progress_callback("File", "status_task_saving_file", 0, os.path.basename(output_path))
                new_workbook.save(output_path)

            else:
                # 일반 모드였으면, 기존 워크북에 직접 수정 후 저장
                workbook_to_save = workbook
                
                # Update cell values first
                for sheet_name in workbook_to_save.sheetnames:
                    sheet = workbook_to_save[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if not hasattr(cell, 'coordinate'): continue
                            if (sheet_name, cell.coordinate) in translated_map: # Use original sheet_name for lookup
                                translated_text = translated_map[(sheet_name, cell.coordinate)]
                                if "오류:" not in translated_text:
                                    cell.value = translated_text
                
                # Then rename sheets
                for sheet_name_original in sheet_name_original_order: # Iterate through original names
                    sheet = workbook_to_save[sheet_name_original] # Get sheet by original name
                    translated_sheet_title = translated_sheet_name_map.get(sheet_name_original, sheet_name_original)
                    sheet.title = translated_sheet_title # Set new title
                
                if progress_callback: progress_callback("File", "status_task_saving_file", 0, os.path.basename(output_path))
                workbook_to_save.save(output_path)

            logger.info(f"번역된 엑셀 파일 저장 완료: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"엑셀 파일 번역 중 최종 오류: {e}", exc_info=True)
            return None
        finally:
            if temp_input_path and os.path.exists(temp_input_path):
                try: os.remove(temp_input_path)
                except OSError as e: logger.warning(f"임시 엑셀 파일 삭제 실패: {e}")

    